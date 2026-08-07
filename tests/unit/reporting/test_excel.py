from openpyxl import load_workbook
import pandas as pd
from credit_risk.reporting.excel import write_excel_report


def test_write_excel_report(tmp_path):

    reports = {
        "00_Summary": pd.DataFrame(
            {
                "metric": ["Loans"],
                "value": [100],
            }
        ),
        "01_Checks": pd.DataFrame(
            {
                "check": ["Grain"],
                "status": ["PASS"],
            }
        ),
    }

    output = tmp_path / "qc.xlsx"

    write_excel_report(
        reports,
        output,
    )

    assert output.exists()

    workbook = load_workbook(output)

    assert workbook.sheetnames == [
        "00_Summary",
        "01_Checks",
    ]
