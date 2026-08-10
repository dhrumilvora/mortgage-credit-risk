from __future__ import annotations

import json
import logging
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import matplotlib.pyplot as plt
from credit_risk.utils.config import create_path

logger = logging.getLogger(__name__)


def _make_json_serializable(value):
    if isinstance(value, pd.DataFrame):
        return value.to_dict(orient="records")

    if isinstance(value, pd.Series):
        return value.to_dict()

    if isinstance(value, np.ndarray):
        return value.tolist()

    if isinstance(value, np.generic):
        return value.item()

    if isinstance(value, dict):
        return {key: _make_json_serializable(val) for key, val in value.items()}

    if isinstance(value, (list, tuple)):
        return [_make_json_serializable(item) for item in value]

    return value


def _write_calibration_sheet(
    workbook,
    calibration,
) -> None:

    ws = workbook.create_sheet("Calibration")

    if calibration is None:
        ws["A1"] = "No calibration results available."
        return

    if isinstance(calibration, pd.DataFrame):
        _write_dataframe(ws, calibration)
    else:
        ws["A1"] = "Unexpected calibration result format."

    _format_sheet(ws)


def _get_evaluation_dir(
    config: dict,
    dataset_name: str,
) -> Path:

    if dataset_name not in {"validation", "oot"}:
        raise ValueError(f"Unsupported evaluation dataset: {dataset_name}")

    evaluation_config = config["parameters"]["evaluation"]

    evaluation_root = create_path(
        config["catalog"]["base"],
        config["catalog"],
        "model_evaluation",
        evaluation_config["model"]["version"],
        evaluation_config["model"]["type"],
        must_exist=False,
    )

    evaluation_dir = evaluation_root / dataset_name
    evaluation_dir.mkdir(
        parents=True,
        exist_ok=True,
    )

    return evaluation_dir


def save_evaluation_json(
    evaluation: dict,
    dataset_name: str,
    config: dict,
) -> None:

    evaluation_dir = _get_evaluation_dir(
        config,
        dataset_name,
    )

    evaluation_path = evaluation_dir / "evaluation_results.json"

    serializable_evaluation = _make_json_serializable(evaluation)

    with open(
        evaluation_path,
        "w",
        encoding="utf-8",
    ) as file:
        json.dump(
            serializable_evaluation,
            file,
            indent=4,
        )

    logger.info(
        "Evaluation results saved: dataset=%s path=%s",
        dataset_name,
        evaluation_path,
    )


def save_evaluation_excel(
    evaluation: dict,
    dataset_name: str,
    config: dict,
) -> None:
    """Save evaluation results as a formatted Excel workbook."""

    evaluation_dir = _get_evaluation_dir(
        config,
        dataset_name,
    )

    evaluation_path = evaluation_dir / "evaluation_report.xlsx"

    workbook = Workbook()

    # Remove default worksheet.
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    _write_summary_sheet(
        workbook,
        evaluation,
        dataset_name,
        config,
    )

    _write_metrics_sheet(
        workbook,
        "Classification Metrics",
        evaluation.get("classification_metrics"),
    )

    _write_metrics_sheet(
        workbook,
        "Risk Metrics",
        evaluation.get("risk_metrics"),
    )

    _write_deciles_sheet(
        workbook,
        evaluation.get("risk_deciles"),
    )

    _write_calibration_sheet(
        workbook,
        evaluation.get("calibration"),
    )

    _write_confusion_matrix_sheet(
        workbook,
        evaluation.get("confusion_matrix"),
    )

    workbook.save(evaluation_path)

    logger.info(
        "Evaluation Excel report saved: dataset=%s path=%s",
        dataset_name,
        evaluation_path,
    )


def _write_summary_sheet(
    workbook,
    evaluation: dict,
    dataset_name: str,
    config: dict,
) -> None:

    ws = workbook.create_sheet("Summary")

    evaluation_config = config["parameters"]["evaluation"]

    rows = [
        ("Dataset", dataset_name),
        ("Model Version", evaluation_config["model"]["version"]),
        ("Model Type", evaluation_config["model"]["type"]),
        (
            "Classification Threshold",
            evaluation_config["classification"]["threshold"],
        ),
    ]

    for row_idx, (label, value) in enumerate(rows, start=1):
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=value)

    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)

    _format_sheet(ws)


def _write_metrics_sheet(
    workbook,
    sheet_name: str,
    metrics: dict | None,
) -> None:

    ws = workbook.create_sheet(sheet_name)

    if not metrics:
        ws["A1"] = "No metrics available."
        return

    ws.append(["Metric", "Value"])

    for key, value in metrics.items():
        ws.append([key, value])

    ws["A1"].font = Font(bold=True)

    _format_sheet(ws)


def _write_deciles_sheet(
    workbook,
    risk_deciles,
) -> None:

    ws = workbook.create_sheet("Risk Deciles")

    if risk_deciles is None:
        ws["A1"] = "No decile results available."
        return

    if isinstance(risk_deciles, pd.DataFrame):
        _write_dataframe(ws, risk_deciles)
    else:
        ws["A1"] = "Unexpected decile result format."

    _format_sheet(ws)


def _write_confusion_matrix_sheet(
    workbook,
    confusion_matrix: dict | None,
) -> None:

    ws = workbook.create_sheet("Confusion Matrix")

    if confusion_matrix is None:
        ws["A1"] = "No confusion matrix available."
        return

    tn = confusion_matrix["true_negative"]
    fp = confusion_matrix["false_positive"]
    fn = confusion_matrix["false_negative"]
    tp = confusion_matrix["true_positive"]

    # ---------------------------------------------------------
    # Counts
    # ---------------------------------------------------------

    ws["A1"] = "Confusion Matrix — Counts"
    ws["A1"].font = Font(bold=True, size=14)

    ws["B2"] = "Predicted 0"
    ws["C2"] = "Predicted 1"

    ws["A3"] = "Actual 0"
    ws["A4"] = "Actual 1"

    ws["B3"] = tn
    ws["C3"] = fp
    ws["B4"] = fn
    ws["C4"] = tp

    for cell in ("B2", "C2", "A3", "A4"):
        ws[cell].font = Font(bold=True)
        ws[cell].alignment = Alignment(horizontal="center")

    for row in range(3, 5):
        for col in range(2, 4):
            ws.cell(
                row=row,
                column=col,
            ).alignment = Alignment(horizontal="center")

    # Heatmap for counts.
    ws.conditional_formatting.add(
        "B3:C4",
        ColorScaleRule(
            start_type="min",
            start_color="FFFFFF",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFF2CC",
            end_type="max",
            end_color="63BE7B",
        ),
    )

    # ---------------------------------------------------------
    # Row-normalized percentages
    # ---------------------------------------------------------

    ws["A7"] = "Confusion Matrix — Row %"
    ws["A7"].font = Font(bold=True, size=14)

    ws["B8"] = "Predicted 0"
    ws["C8"] = "Predicted 1"

    ws["A9"] = "Actual 0"
    ws["A10"] = "Actual 1"

    actual_0 = tn + fp
    actual_1 = fn + tp

    if actual_0 > 0:
        ws["B9"] = tn / actual_0
        ws["C9"] = fp / actual_0
    else:
        ws["B9"] = 0
        ws["C9"] = 0

    if actual_1 > 0:
        ws["B10"] = fn / actual_1
        ws["C10"] = tp / actual_1
    else:
        ws["B10"] = 0
        ws["C10"] = 0

    for row in range(9, 11):
        for col in range(2, 4):
            cell = ws.cell(
                row=row,
                column=col,
            )
            cell.number_format = "0.0%"
            cell.alignment = Alignment(horizontal="center")

    ws.conditional_formatting.add(
        "B9:C10",
        ColorScaleRule(
            start_type="min",
            start_color="FFFFFF",
            mid_type="percentile",
            mid_value=50,
            mid_color="FFF2CC",
            end_type="max",
            end_color="63BE7B",
        ),
    )

    _format_sheet(ws)


def _write_dataframe(
    worksheet,
    dataframe: pd.DataFrame,
) -> None:

    for column_idx, column in enumerate(
        dataframe.columns,
        start=1,
    ):
        worksheet.cell(
            row=1,
            column=column_idx,
            value=column,
        )

    for row_idx, row in enumerate(
        dataframe.itertuples(index=False),
        start=2,
    ):
        for column_idx, value in enumerate(
            row,
            start=1,
        ):
            worksheet.cell(
                row=row_idx,
                column=column_idx,
                value=value,
            )


def _format_sheet(worksheet) -> None:

    worksheet.freeze_panes = "A2"

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for column_cells in worksheet.columns:
        max_length = 0

        for cell in column_cells:
            if cell.value is not None:
                max_length = max(
                    max_length,
                    len(str(cell.value)),
                )

        column_letter = get_column_letter(column_cells[0].column)

        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 40)


def save_evaluation_charts(
    evaluation: dict,
    dataset_name: str,
    config: dict,
) -> None:

    evaluation_dir = _get_evaluation_dir(
        config,
        dataset_name,
    )

    _save_roc_curve(
        evaluation,
        evaluation_dir,
        dataset_name,
    )

    _save_ks_curve(
        evaluation,
        evaluation_dir,
        dataset_name,
    )

    _save_risk_decile_chart(
        evaluation,
        evaluation_dir,
        dataset_name,
    )

    _save_calibration_chart(
        evaluation,
        evaluation_dir,
        dataset_name,
    )


def _save_roc_curve(
    evaluation: dict,
    evaluation_dir: Path,
    dataset_name: str,
) -> None:

    roc_data = evaluation["roc_curve"]

    fpr = np.asarray(roc_data["fpr"])
    tpr = np.asarray(roc_data["tpr"])

    roc_auc = evaluation["ds_metrics"]["roc_auc"]

    fig, ax = plt.subplots(figsize=(8, 6))

    ax.plot(
        fpr,
        tpr,
        label=f"ROC AUC = {roc_auc:.4f}",
    )

    ax.plot(
        [0, 1],
        [0, 1],
        linestyle="--",
        label="Random",
    )

    ax.set_title(f"ROC Curve — {dataset_name.upper()}")
    ax.set_xlabel("False Positive Rate")
    ax.set_ylabel("True Positive Rate")

    ax.legend()
    ax.grid(alpha=0.3)

    fig.tight_layout()

    fig.savefig(
        evaluation_dir / "roc.png",
        dpi=200,
        bbox_inches="tight",
    )

    plt.close(fig)


def _save_ks_curve(
    evaluation: dict,
    evaluation_dir: Path,
    dataset_name: str,
) -> None:

    ks_data = evaluation["ks_curve"]

    population_pct = ks_data["population_pct"]
    cum_bad = ks_data["cum_bad"]
    cum_good = ks_data["cum_good"]
    ks = ks_data["ks"]

    max_ks_idx = np.argmax(ks)

    fig, ax = plt.subplots(figsize=(8, 6))

    ax.plot(
        population_pct,
        cum_bad,
        label="Cumulative Bad",
    )

    ax.plot(
        population_pct,
        cum_good,
        label="Cumulative Good",
    )

    ax.axvline(
        population_pct.iloc[max_ks_idx],
        linestyle="--",
        label=f"KS = {ks.iloc[max_ks_idx]:.4f}",
    )

    ax.set_title(f"KS Curve — {dataset_name.upper()}")
    ax.set_xlabel("Population Proportion")
    ax.set_ylabel("Cumulative Distribution")

    ax.legend()
    ax.grid(alpha=0.3)

    fig.tight_layout()

    fig.savefig(
        evaluation_dir / "ks.png",
        dpi=200,
        bbox_inches="tight",
    )

    plt.close(fig)


def _save_risk_decile_chart(
    evaluation: dict,
    evaluation_dir: Path,
    dataset_name: str,
) -> None:

    deciles = evaluation["risk_deciles"]

    fig, ax = plt.subplots(figsize=(8, 6))

    ax.plot(
        deciles["risk_decile"],
        deciles["actual_event_rate"],
        marker="o",
        label="Observed Event Rate",
    )

    ax.plot(
        deciles["risk_decile"],
        deciles["average_predicted_df"],
        marker="o",
        label="Average Predicted PD",
    )

    ax.set_title(f"Risk Deciles — {dataset_name.upper()}")
    ax.set_xlabel("Risk Decile")
    ax.set_ylabel("Probability of Default")

    ax.legend()
    ax.grid(alpha=0.3)

    fig.tight_layout()

    fig.savefig(
        evaluation_dir / "risk_deciles.png",
        dpi=200,
        bbox_inches="tight",
    )

    plt.close(fig)


def _save_calibration_chart(
    evaluation: dict,
    evaluation_dir: Path,
    dataset_name: str,
) -> None:

    calibration = evaluation["calibration"]

    fig, ax = plt.subplots(figsize=(8, 6))

    ax.plot(
        calibration["average_predicted_pd"],
        calibration["actual_event_rate"],
        marker="o",
        label="Observed",
    )

    ax.plot(
        [0, 1],
        [0, 1],
        linestyle="--",
        label="Perfect Calibration",
    )

    ax.set_title(f"Calibration — {dataset_name.upper()}")
    ax.set_xlabel("Average Predicted PD")
    ax.set_ylabel("Observed Event Rate")

    ax.legend()
    ax.grid(alpha=0.3)

    fig.tight_layout()

    fig.savefig(
        evaluation_dir / "calibration.png",
        dpi=200,
        bbox_inches="tight",
    )

    plt.close(fig)


def save_evaluation_results(
    validation_evaluation: dict | None,
    oot_evaluation: dict | None,
    config: dict,
) -> None:
    """
    Save all evaluation artifacts for available datasets.

    Generates JSON, Excel, and chart outputs for each
    evaluated dataset.
    """

    evaluations = {
        "validation": validation_evaluation,
        "oot": oot_evaluation,
    }

    for dataset_name, evaluation in evaluations.items():

        if evaluation is None:
            logger.info(
                "Skipping evaluation reporting: dataset=%s",
                dataset_name,
            )
            continue

        logger.info(
            "Saving evaluation artifacts: dataset=%s",
            dataset_name,
        )

        save_evaluation_json(
            evaluation=evaluation,
            dataset_name=dataset_name,
            config=config,
        )

        save_evaluation_excel(
            evaluation=evaluation,
            dataset_name=dataset_name,
            config=config,
        )

        save_evaluation_charts(
            evaluation=evaluation,
            dataset_name=dataset_name,
            config=config,
        )

    logger.info("Evaluation artifact generation completed.")
