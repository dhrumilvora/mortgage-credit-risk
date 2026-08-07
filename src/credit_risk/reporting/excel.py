"""Excel writer for data-quality reports."""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = "1F4E78"
HEADER_FONT = "FFFFFF"

PASS_FILL = "C6EFCE"
PASS_FONT = "006100"

WARN_FILL = "FFEB9C"
WARN_FONT = "9C6500"

FAIL_FILL = "FFC7CE"
FAIL_FONT = "9C0006"

INFO_FILL = "D9EAF7"


def write_dataframe(
    worksheet,
    df: pd.DataFrame,
) -> None:
    """Write a DataFrame to an Excel worksheet."""

    # Header
    for col_idx, column in enumerate(df.columns, start=1):
        cell = worksheet.cell(
            row=1,
            column=col_idx,
            value=column,
        )

        cell.fill = PatternFill(
            "solid",
            fgColor=HEADER_FILL,
        )

        cell.font = Font(
            color=HEADER_FONT,
            bold=True,
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    # Data
    for row_idx, row in enumerate(
        df.itertuples(index=False, name=None),
        start=2,
    ):
        for col_idx, value in enumerate(row, start=1):

            # Convert pandas/numpy missing values to Excel blank
            if pd.isna(value):
                value = None

            worksheet.cell(
                row=row_idx,
                column=col_idx,
                value=value,
            )


def format_status_column(
    worksheet,
    df: pd.DataFrame,
) -> None:
    """Apply PASS/WARN/FAIL/INFO formatting."""

    if "status" not in df.columns:
        return

    status_col = df.columns.get_loc("status") + 1

    for row in range(2, len(df) + 2):
        cell = worksheet.cell(
            row=row,
            column=status_col,
        )

        status = cell.value

        if status == "PASS":
            cell.fill = PatternFill(
                "solid",
                fgColor=PASS_FILL,
            )
            cell.font = Font(
                color=PASS_FONT,
                bold=True,
            )

        elif status == "WARN":
            cell.fill = PatternFill(
                "solid",
                fgColor=WARN_FILL,
            )
            cell.font = Font(
                color=WARN_FONT,
                bold=True,
            )

        elif status == "FAIL":
            cell.fill = PatternFill(
                "solid",
                fgColor=FAIL_FILL,
            )
            cell.font = Font(
                color=FAIL_FONT,
                bold=True,
            )

        elif status == "INFO":
            cell.fill = PatternFill(
                "solid",
                fgColor=INFO_FILL,
            )


def format_percentage_columns(
    worksheet,
    df: pd.DataFrame,
) -> None:
    """Format percentage/rate columns."""

    percentage_keywords = (
        "pct",
        "rate",
        "share",
        "retention",
    )

    for col_idx, column in enumerate(df.columns, start=1):

        column_lower = column.lower()

        if any(keyword in column_lower for keyword in percentage_keywords):
            for row in range(2, len(df) + 2):
                worksheet.cell(
                    row=row,
                    column=col_idx,
                ).number_format = "0.00%"


def resize_columns(
    worksheet,
    max_width: int = 40,
) -> None:
    """Resize worksheet columns based on displayed content."""

    for column_cells in worksheet.columns:

        column_letter = get_column_letter(column_cells[0].column)

        max_length = 0

        for cell in column_cells:
            if cell.value is not None:
                max_length = max(
                    max_length,
                    len(str(cell.value)),
                )

        worksheet.column_dimensions[column_letter].width = min(
            max(max_length + 2, 10),
            max_width,
        )


def format_worksheet(
    worksheet,
    df: pd.DataFrame,
) -> None:
    """Apply standard formatting to a report worksheet."""

    worksheet.freeze_panes = "A2"

    if worksheet.max_row > 1:
        worksheet.auto_filter.ref = worksheet.dimensions

    worksheet.row_dimensions[1].height = 22

    format_status_column(
        worksheet,
        df,
    )

    format_percentage_columns(
        worksheet,
        df,
    )

    resize_columns(
        worksheet,
    )


def write_excel_report(
    reports: dict[str, pd.DataFrame],
    output_path: str | Path,
) -> Path:
    """
    Write the complete data-quality report to Excel.

    Each report DataFrame is written to a separate worksheet.
    """

    output_path = Path(output_path)

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    workbook = Workbook()

    # Remove default worksheet.
    workbook.remove(workbook.active)

    for sheet_name, df in reports.items():

        # Excel sheet names cannot exceed 31 characters.
        sheet_name = sheet_name[:31]

        worksheet = workbook.create_sheet(title=sheet_name)

        write_dataframe(
            worksheet,
            df,
        )

        format_worksheet(
            worksheet,
            df,
        )

    workbook.save(output_path)

    return output_path
